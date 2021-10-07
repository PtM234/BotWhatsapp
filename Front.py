import random
import threading
import lecturaExcel
import copytoClipboard
import time
import tkinter as tk
from datetime import date, datetime
import pyautogui as pg
import win32api
import webbrowser as web
from tkinter import ttk
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import askdirectory
from urllib.parse import quote
from io import StringIO
from PIL import Image
import win32clipboard
from openpyxl import load_workbook, Workbook


class App(tk.Tk):
    def __init__(self):
        #Funciones que necesito:
        #1.- Leer Excel de recipientes{
        #       List donde guardar los recipientes
        #}
        #2.- Leer Excel de mensajes{
        #       List donde guardar los mensajes
        #}
        #3.- Especificar directorio para la impresión del reporte{
        #       List donde se guarde qué mensaje se le manda a qué numero
        #}
        #4.- Copiar imagen a clipboard
        #5.-

        super().__init__()
        opts = {'ipadx': 50, 'ipady': 5, 'fill': tk.Y}
        btn_opts = {"expand": True, "fill": tk.X, "ipadx": 50, "padx": 20, "pady": 5}
        #ipadx es para separar las letras DENTRO del botón, padx es para separar el botón de los bordes de la interfaz

        self.listaContactos = []
        self.listaMensajes = []
        self.listaAcciones = []
        self.direccionDirectorio = ''
        self.direccionImagen = ''
        self.boolDireccionDirectorios = False
        self.boolContactos = False
        self.boolMensajes = False
        self.Running = False
        self.boolImg = False
        self.browser = None
        self.nombreArchivo = ''
        self.contador = 1
        self.longitudContactos = 0
        self.opciones = [15, 20, 30, 45, 60]

        self.listaDesplegable = ttk.Combobox(self)
        self.listaDesplegable['values'] = self.opciones
        self.labelTiempo = tk.Label(self, text="Por favor indica cada cuanto quieres que se ejecute una acción")
        self.btnMensajes = tk.Button(self, text="Elegir mensajes a enviar", command=self.obtenerListaMensajes)
        self.btnContactos = tk.Button(self, text="Elegir lista de destinatarios", command=self.obtenerListaContactos)
        self.btnImagen = tk.Button(self, text="Elegir imagen a enviar", command=self.copiarAClipboard)
        self.btnComenzar = tk.Button(self, text="Comenzar ejecución", command=self.comenzar)
        self.btnDetener = tk.Button(self, text="Detener", command=self.terminoEjecucion)
        self.btnDirectorio = tk.Button(self, text="Elige un directorio donde guardar los reportes", command=self.obtenerDirectorio)

        self.btnContactos.pack(**btn_opts)
        self.btnMensajes.pack(**btn_opts)
        self.btnImagen.pack(**btn_opts)
        self.btnDirectorio.pack(**btn_opts)
        self.btnComenzar.pack(**btn_opts)
        self.btnDetener.pack(**btn_opts)
        self.labelTiempo.pack(**opts)
        self.listaDesplegable.pack(**btn_opts)


    def fechaActual(self):
        now = datetime.now()
        hora = now.strftime('%H%M%S')
        day = now.strftime("%d")
        month = now.strftime("%B")
        year = now.strftime("%Y")
        self.nombreArchivo = self.direccionDirectorio + f'/Acciones realizadas el día {day} de {month} del {year} a las {hora}.xlsx'
        print(self.nombreArchivo)

    def comenzar(self):
        if self.boolDireccionDirectorios:
            if self.boolContactos:
                if self.boolMensajes:
                    self.Running = True
                    self.fechaActual()
                    if self.boolImg:
                        self.fechaActual()
                        self.desactivarBotones()
                        t1 = threading.Thread(target = self.enviarMsgsImg())
                        t1.daemon = True
                        t1.start()
                    else:
                        self.fechaActual()
                        self.desactivarBotones()
                        t1 = threading.Thread(target = self.enviarMsgs())
                        t1.daemon = True
                        t1.start()
                elif not self.boolMensajes:
                    win32api.MessageBox(0, '¡Elige un archivo que contenga los mensajes!', 'Error', 0x00001000)
            elif not self.boolContactos:
                win32api.MessageBox(0, '¡Elige un archivo que contenga los recipientes!', 'Error', 0x00001000)
        elif not self.boolDireccionDirectorios:
            win32api.MessageBox(0, '¡Elige un directorio donde guardar los reportes!', 'Error', 0x00001000)

    def detener(self):
        print("Se detuvo el programa")
        self.terminoEjecucion()
        self.imprimirAccionesExcel()
        self.imprimirFaltantes()
        self.activarBotones()

    def enviarMsgsImg(self):
        if self.Running:
            if self.browser and self.browser.lower() not in ["chrome", "firefox", "brave", "opera"]:
                win32api.MessageBox(0, "Solo funciona con Chrome, Firefox, Brave y Opera", "Error")
                self.Running = False
            elif self.browser:
                parsedMessage = quote(self.obtenerMensaje())
                numeroTelefono = self.obtenerNumero()
                web.open('https://web.whatsapp.com/send?phone=' + numeroTelefono + '&text=' + parsedMessage)
                whats = pg.getWindowsWithTitle(self.browser)[0]
                whats.maximize()
                whats.activate()
                width, height = pg.size()
                pg.click(width / 2, height / 2)
                pg.hotkey('ctrl', 'v')
                time.sleep(15)
                pg.press('enter')
                self.listaAcciones.append(f"Mensaje: {parsedMessage} enviado al numero: {numeroTelefono}")
                print(f"Lugar #{self.contador} de {len(self.longitudContactos)}")
                if len(self.listaContactos>=1):
                    tempo = int(self.listaDesplegable.get())
                    time.sleep(tempo + random.randint(10, 25))
                    self.enviarMsgsImg()
                else:
                    win32api.MessageBox(0, "Se terminó la ejecución del programa", "Terminado")
                    self.imprimirAccionesExcel()

    def enviarMsgs(self):
        if self.Running:
            if self.browser and self.browser.lower() not in ["chrome", "firefox", "brave", "opera"]:
                win32api.MessageBox(0, "Solo funciona con Chrome, Firefox, Brave y Opera", "Error")
                self.Running = False
            elif self.browser:
                parsedMessage = quote(self.obtenerMensaje())
                numeroTelefono = self.obtenerNumero()
                web.open('https://web.whatsapp.com/send?phone=' + numeroTelefono + '&text=' + parsedMessage)
                whats = pg.getWindowsWithTitle(self.browser)[0]
                whats.maximize()
                whats.activate()
                width, height = pg.size()
                pg.click(width / 2, height / 2)
                pg.press('enter')
                self.listaAcciones.append(f"Mensaje: {parsedMessage} enviado al numero: {numeroTelefono}")
                print(f"Lugar #{self.contador} de {len(self.longitudContactos)}")
                if len(self.listaContactos>=1):
                    tempo = int(self.listaDesplegable.get())
                    time.sleep(tempo + random.randint(10, 25))
                    self.enviarMsgs()
                else:
                    win32api.MessageBox(0, "Se terminó la ejecución del programa", "Terminado")
                    self.imprimirAccionesExcel()

    def obtenerDirectorio(self):
        try:
            self.direccionDirectorio = askdirectory(title='Escoge el directorio donde guardar los reportes')  # shows dialog box and return the path
            self.boolDireccionDirectorios = True
            print(self.direccionDirectorio)
        except:
            print("Ocurrió un error al elegir el directorio donde se guardarán los reportes")

    def obtenerMensaje(self):
        if len(self.listaMensajes) > 1:
            numeroMensaje = random.randint(0, len(self.listaMensajes) - 1)
            mensaje = self.listaMensajes[numeroMensaje]
            return mensaje
        else:
            mensaje = self.listaMensajes[0]
            return mensaje

    def obtenerNumero(self):
        numero = self.listaContactos[0]
        self.listaContactos.pop(0)

    def copiarAClipboard(self):
        try:
            self.direccionImagen = askopenfilename()
            imagen = Image.open(self.direccionImagen)
            output = StringIO()
            imagen.convert("RGB").save(output, "BMP")
            data = output.getvalue()[14:]
            output.close()
            copytoClipboard.copiarAClipboard(win32clipboard.CF_DIB, data)
            self.boolImg = True
        except:
            print("Error abriendo archivo")

    def obtenerListaContactos(self):
        self.listaContactos = lecturaExcel.listaRecipientes()
        self.boolContactos = True
        self.longitudContactos = len(self.listaContactos)

    def obtenerListaMensajes(self):
        self.listaMensajes = lecturaExcel.listaMensajes()
        self.boolMensajes = True

    def imprimirAccionesExcel(self):
        if self.direccionDirectorio:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Accion'
            r = 2
            for accion in self.listaAcciones:
                ws.cell(row = r, column = 1).value = accion
                r = r+1
            wb.save(self.nombreArchivo)
            wb.close()

    def imprimirFaltantes(self):
        if self.direccionDirectorio:
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Accion'
            r = 2
            for contacto in self.listaContactos:
                ws.cell(row = r, column = 1).value = contacto
                r = r+1
            wb.save("numerosFaltantes.xlsx")
            wb.close()

    def terminoEjecucion(self):
        self.Running = False
        self.boolImg = False
        self.boolMensajes = False
        self.boolContactos = False
        self.boolDireccionDirectorios = False
        self.activarBotones()

    def desactivarBotones(self):
        self.btnComenzar.config(state = tk.DISABLED)
        self.btnDirectorio.config(state = tk.DISABLED)
        self.btnImagen.config(state = tk.DISABLED)
        self.btnMensajes.config(state = tk.DISABLED)
        self.btnContactos.config(state = tk.DISABLED)

    def activarBotones(self):
        self.btnDirectorio.config(state = tk.NORMAL)
        self.btnImagen.config(state = tk.NORMAL)
        self.btnMensajes.config(state = tk.NORMAL)
        self.btnContactos.config(state = tk.NORMAL)


if __name__ == "__main__":
    app = App()
    app.mainloop()

