from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename

def listaRecipientes():
    recipientes = []
    archivoALeerRecipientes = askopenfilename()
    recipientesExcel=load_workbook(filename=archivoALeerRecipientes)
    sheet = recipientesExcel.active
    list = sheet["A"]
    for cel in list:
        recipientes.append(str(cel.value))
    return recipientes


def listaMensajes():
    mensajes = []
    archivoALeerMensajes = askopenfilename()
    mensajesExcel = load_workbook(filename=archivoALeerMensajes)
    sheet = mensajesExcel.active
    list = sheet["A"]
    for mensaje in list:
        mensajes.append(str(mensaje.value))
    return mensajes